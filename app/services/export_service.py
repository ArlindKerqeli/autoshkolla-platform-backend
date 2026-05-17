"""Shared utility for Excel/CSV export generation."""

import csv
import io
from dataclasses import dataclass
from typing import Any, Callable, Optional
from datetime import datetime, date

from flask import make_response, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


@dataclass
class ExportColumn:
    """Defines a column in an export file."""
    key: str              # dict key to read from row
    header: str           # Albanian column header
    width: int = 20       # column width
    formatter: Optional[Callable[[Any], Any]] = None  # optional value transformer


def format_date(val: Any) -> str:
    """Format date/datetime to dd.MM.yyyy"""
    if isinstance(val, datetime):
        return val.strftime('%d.%m.%Y')
    if isinstance(val, date):
        return val.strftime('%d.%m.%Y')
    if isinstance(val, str) and val:
        # Try parsing ISO format
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
            try:
                return datetime.strptime(val[:19], fmt).strftime('%d.%m.%Y')
            except ValueError:
                continue
    return val or ''


def format_currency(val: Any) -> str:
    """Format as Euro currency."""
    if val is None:
        return '0.00 \u20ac'
    try:
        return f'{float(val):.2f} \u20ac'
    except (ValueError, TypeError):
        return '0.00 \u20ac'


def format_boolean(val: Any) -> str:
    """Format boolean as Po/Jo (Albanian Yes/No)."""
    return 'Po' if val else 'Jo'


def _get_cell_value(row: dict, col: ExportColumn) -> Any:
    """Extract and optionally format a cell value from a row dict."""
    val = row.get(col.key, '')
    if col.formatter:
        return col.formatter(val)
    return val if val is not None else ''


def generate_excel(rows: list[dict], columns: list[ExportColumn], sheet_name: str = 'Data') -> bytes:
    """Generate an Excel file from row data and column definitions."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

    # Write headers
    for col_idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left')
        ws.column_dimensions[cell.column_letter].width = col.width

    # Write data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(columns, 1):
            val = _get_cell_value(row, col)
            ws.cell(row=row_idx, column=col_idx, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_csv(rows: list[dict], columns: list[ExportColumn]) -> bytes:
    """Generate a CSV file from row data and column definitions."""
    buf = io.StringIO()
    # UTF-8 BOM for proper Albanian character display in Excel
    buf.write('\ufeff')
    writer = csv.writer(buf)

    # Write headers
    writer.writerow([col.header for col in columns])

    # Write data rows
    for row in rows:
        writer.writerow([_get_cell_value(row, col) for col in columns])

    return buf.getvalue().encode('utf-8')


def make_export_response(content: bytes, filename: str, fmt: str) -> Response:
    """Create a Flask response for file download."""
    if fmt == 'csv':
        content_type = 'text/csv; charset=utf-8'
        ext = 'csv'
    else:
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ext = 'xlsx'

    if not filename.endswith(f'.{ext}'):
        filename = f'{filename}.{ext}'

    response = make_response(content)
    response.headers['Content-Type'] = content_type
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
